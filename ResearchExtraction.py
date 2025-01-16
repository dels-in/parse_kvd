import os
import re
import subprocess
from datetime import datetime, timedelta

import pandas as pd
import textract
from docx import Document
from openpyxl.styles import Border, Side

research_headers = ['Скважина',
                    'Дата, приведенная на конец исследования',
                    'Начало исследования',
                    'Конец исследования',
                    'Время исследования, ч',
                    'Глубина прибора (мандрель), м',
                    'СИП, м',
                    'ВДП, м',
                    'Средняя плотность флюида в стволе скважины, г/см3',
                    'Замеренное давление на первую точку КВД на глубине мандрели, МПа а',
                    'Замеренное давление на последнюю точку КВД на глубине мандрели, МПа а',
                    'Забойное давление на последнюю точку КВД на глубине ВДП, МПа а',
                    'Расчетное пластовое давление в области дренирования на глубине мандрели, МПа а',
                    'Пластовое давление в области дренирования на глубине СИП, МПа а',
                    'Среднее расчетное пластовое давление в области дренирования на глубине ВДП, МПа а']

# ANSI escape code для красного фона
red_background = "\033[41m"
# ANSI escape code для сброса стиля
reset = "\033[0m"


def extract_info_from_table(table):
    data = []
    for row in table.rows:
        if len(row.cells) == 2:
            cell1 = row.cells[0].text
            cell2 = row.cells[1].text
            data.append([cell1, cell2])
    return data


def dance(i, file_name):
    if i % 4 == 0:
        print(f'{file_name} в работе ヽ(￣▽￣)ﾉ')
    elif i % 4 == 1:
        print(f'{file_name} в работе \(▔∀▔)/')
    elif i % 4 == 2:
        print(f'{file_name} в работе ╰(￣▽￣)ノ')
    else:
        print(f'{file_name} в работе \(▔∀▔)/')
    i += 1
    return i


def find_pressure(text, file_path):
    text = text.replace(' ', '').replace('\n', '')
    pattern = r"Забойноедавлениенациклевосстановленияизменилосьот(\d+([.,]\d+)?)кгс/см2до(\d+([.,]\d+)?)кгс/см2"
    match = re.search(pattern, text)
    if match:
        return match.group(1), match.group(3)
    else:
        text = read_doc_with_libreoffice(file_path).replace(' ', '').replace('\n', '')
        match = re.search(pattern, text)
        if match:
            return match.group(1), match.group(3)
        else:
            print(f"{red_background}Замеренные забойные давления не найдены.{reset}")
            return None, None


def fill_empty_first_values(data_list):
    previous_first_value = None
    for i in range(len(data_list)):
        if data_list[i][0].strip() == '':
            if previous_first_value is not None:
                data_list[i][0] = previous_first_value
        else:
            previous_first_value = data_list[i][0]
    return data_list


def convert_doc_to_text(file_path, output_path):
    # Вызываем LibreOffice для преобразования файла .doc в текстовый формат
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'txt', '--outdir', output_path, file_path],
        capture_output=True,
        text=True
    )

    # Проверяем, успешно ли выполнена команда
    if result.returncode == 0:
        return True
    else:
        print(f"Error: {result.stderr}")
        return False


def read_doc_with_libreoffice(file_path):
    # Определяем путь для выходного файла
    output_path = os.path.dirname(file_path)
    output_file = os.path.join(output_path, os.path.basename(file_path).replace('.doc', '.txt'))

    # Преобразуем файл .doc в текстовый формат
    if convert_doc_to_text(file_path, output_path):
        # Читаем текст из выходного файла
        with open(output_file, 'r', encoding='utf-8') as f:
            text = f.read()

        # Удаляем файл .txt после чтения
        os.remove(output_file)

        return text
    else:
        return None



def get_from_doc(file_path):
    text = textract.process(file_path)
    data = text.decode()
    downhole_start, downhole_end = find_pressure(data, file_path)
    data = data.lower().replace('расчетное', '').replace('текущее', '')
    data_list = [line.strip().strip('|').split('|') for line in data.split('\n')
                 if '|' in line.strip('|') and len(line.strip().strip('|').split('|')) > 1]
    data_list = fill_empty_first_values(data_list)

    data_dict = {}
    for line in data_list:
        key = clean_data(''.join(line[0].strip().split()))
        value = clean_data(''.join(line[1].strip().split()), False)
        if key in data_dict:
            data_dict[key].append(value)
        else:
            data_dict[key] = [value]

    data_df = pd.DataFrame([data_dict])
    return data_df, downhole_start, downhole_end


def get_from_docx(file_path):
    doc = Document(file_path)
    data = []
    for table in doc.tables:
        data.extend(extract_info_from_table(table))

    data_df = pd.DataFrame(data, columns=['data'])
    return data_df


def get_from_excel(file_path):
    if file_path.endswith('.xls'):
        # Чтение файла .xls с использованием xlrd
        excel = pd.ExcelFile(file_path, engine='xlrd')
    elif file_path.endswith('.xlsx'):
        # Чтение файла .xlsx с использованием openpyxl
        excel = pd.ExcelFile(file_path, engine='openpyxl')

    for sheet_name in excel.sheet_names:
        data_df = pd.read_excel(excel, sheet_name=sheet_name)
        data_df = data_df.dropna(axis=1, how='all')
    return data_df


# def get_from_pdf(file_path):
#     det_models, rec_models, layout_models = load_detection_models(), load_recognition_models(), load_layout_models()
#     images, highres_images, names, text_lines = load_pdfs_images(file_path)
#
#     page_results = extract_tables(images, highres_images, text_lines, det_models, layout_models, rec_models)
#     data_df = pd.DataFrame()
#     return data_df


def clean_data(value, col=True):
    """ Функция для удаления запятых и данных между скобками """
    if isinstance(value, str):
        # Если название параметра, производим очистку
        if col:
            # Удаляем данные между скобками
            value = re.sub(r'\(кгс/см2\)', '', value)
            value = re.sub(r'ач.*?,', ',', value)

            value = value.replace(',', '')
            value = value.replace('(', '')
            value = value.replace(')', '')
        # Если сам параметр, меняем "," на ".", чтобы корректно обрабатывать числа с плавающей запятой
        else:
            value = value.replace(',', '.')
    return value


def calculate_average(value):
    if '-' in value:
        start, end = map(float, value.split('-'))
        return (start + end) / 2
    else:
        return float(value)


def extract_perforation_interval(values):
    numbers = []
    for value in values:
        if any(char.isalpha() for char in value):
            # Используем регулярное выражение для извлечения чисел до первого буквенного символа
            match = re.match(r'(\d+\.\d+|\d+)', value)
            if match:
                numbers.append(match.group(1))
        else:
            # Если строка содержит только числа, разделенные дефисом
            numbers.extend(re.findall(r'\d+\.\d+|\d+', value))
    if numbers:
        first_number = float(min(numbers, key=lambda x: float(x)))
        last_number = float(max(numbers, key=lambda x: float(x)))
        return first_number, (first_number + last_number) / 2
    else:
        return None, None


def get_value_or_none(data, column_name):
    try:
        return data[column_name].iloc[0]
    except KeyError:
        print(f"{red_background}Error: Column '{column_name}' is missing in the DataFrame.{reset}")
        return None


def reformat_data_headers(data, well_number, downhole_start, downhole_end):
    reduced_date = None

    try:
        date_start = datetime.strptime(data['датаисследования'].iloc[0][0], "%d.%m.%Y")
    except KeyError:
        print(f"{red_background}Ошибка: Столбец 'датаисследования' отсутствует в DataFrame.{reset}")
        date_start = None

    try:
        research_time = timedelta(hours=float(data['общеевремяисследованиячас'].iloc[0][0]))
    except KeyError:
        print(f"{red_background}Ошибка: Столбец 'общеевремяисследованиячас' отсутствует в DataFrame.{reset}")
        research_time = None

    try:
        date_end = date_start + research_time if date_start and research_time else None
    except TypeError:
        print(f"{red_background}Ошибка: Невозможно добавить timedelta к NoneType.{reset}")
        date_end = None

    try:
        depth = float(data['глубинаустановкидатчикам'].iloc[0][0])
    except KeyError:
        print(f"{red_background}Ошибка: Столбец 'глубинаустановкидатчикам' отсутствует в DataFrame.{reset}")
        depth = None

    try:
        VDP, SIP = extract_perforation_interval(data['интервалперфорациим'].iloc[0])
    except KeyError:
        print(f"{red_background}Ошибка: Столбец 'интервалперфорациим' отсутствует в DataFrame.{reset}")
        VDP, SIP = None, None

    density = None
    measured_pressure_VDP = None  # data['забойноедавлениенавдппласта'] - надо уточнить

    try:
        mean_pressure_mandrel = calculate_average(data['пластовоедавлениенаглубинезамера'].iloc[0][0])
    except KeyError:
        print(f"{red_background}Ошибка: Столбец 'пластовоедавлениенаглубинезамера' отсутствует в DataFrame.{reset}")
        mean_pressure_mandrel = None

    SIP_pressure = None

    try:
        mean_pressure_VDP = calculate_average(data['пластовоедавлениенавдппласта'].iloc[0][0])
    except KeyError:
        print(f"{red_background}Ошибка: Столбец 'пластовоедавлениенавдппласта' отсутствует в DataFrame.{reset}")
        mean_pressure_VDP = None

    reformatted_data = pd.DataFrame(
        [[well_number,
          reduced_date,
          date_start.date() if date_start else None,
          date_end.date() if date_end else None,
          (research_time.seconds + research_time.microseconds / 1000) / 3600 + research_time.days * 24 if research_time else None,
          depth,
          SIP,
          VDP,
          density,
          float(downhole_start) * 0.09806650125 if downhole_start is not None else None,
          float(downhole_end) * 0.09806650125 if downhole_end is not None else None,
          measured_pressure_VDP * 0.09806650125 if measured_pressure_VDP is not None else None,
          mean_pressure_mandrel * 0.09806650125 if mean_pressure_mandrel is not None else None,
          SIP_pressure * 0.09806650125 if SIP_pressure is not None else None,
          mean_pressure_VDP * 0.09806650125 if mean_pressure_VDP is not None else None]],
        columns=research_headers)

    return reformatted_data


def extract_researches(file_path, well_number):
    data = []
    if file_path.endswith('.doc'):
        data, downhole_start, downhole_end = get_from_doc(file_path)
        data = reformat_data_headers(data, well_number, downhole_start, downhole_end)
    elif file_path.endswith('.docx'):
        data = get_from_docx(file_path)
    # elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
    #     data = get_from_excel(file_path)
    # elif file_path.endswith('.pdf'):
    #     data = get_from_pdf(file_path)
    else:
        print(file_path + ' не обработан')

    return data


def prepare_dfs_to_write(dfs):
    concatenated_df = pd.concat(dfs, ignore_index=True)
    sorted_df = concatenated_df.sort_values(by='Начало исследования')
    return sorted_df


def write_to_excel(combined_dfs, output_path, ):
    print(f'(ﾉ◕ヮ◕)ﾉ*:･ﾟ✧ Сохранение в {output_path}')

    df_to_write = prepare_dfs_to_write(combined_dfs)

    # Проверяем, существует ли файл
    if os.path.exists(output_path):
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_to_write.to_excel(writer, index=False)

            # Добавляем границы к ячейкам
            wb = writer.book

            ws = wb.create_sheet(title='Лист 1')

            # Определяем стиль границы
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Применяем границы ко всем ячейкам
            for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.border = border_style
    else:
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
            df_to_write.to_excel(writer, index=False)

            # Добавляем границы к ячейкам
            wb = writer.book

            ws = wb.create_sheet(title='Лист 1')

            # Определяем стиль границы
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Применяем границы ко всем ячейкам
            for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.border = border_style

    print(f'Данные Pпл успешно объединены и сохранены в {output_path} (♡°▽°♡)')


def find_well_folder(root_path, current_path):
    current_dir = current_path
    while current_dir != root_path:
        folder_name = os.path.basename(current_dir)
        match = re.search(r'\d+[АAА]\d+', folder_name, re.IGNORECASE)
        if match:
            return match.group()
        current_dir = os.path.dirname(current_dir)
    return None


def get_area_number(file_path):
    # Ищем совпадение с номером участка в формате /{Численный номер участка}{Буква "А" или "A"}
    match = re.search(r'/(\d+)[АA]/', file_path)
    if match:
        area_number = match.group(1) + 'A'
        return area_number
    else:
        return None


def process_files_in_directory(folder_path, output_folder_path=os.getcwd()):
    all_researches = []

    for root, dirs, files in os.walk(folder_path):
        well_folder_name = find_well_folder(folder_path, root)  # ищем номер скважины
        i = 0
        if well_folder_name:
            if "ГКИ" in root:
                for file_name in files:

                    file_path = os.path.join(root, file_name)
                    if '~$' in file_path or '.~' in file_path or '!Нету_' in file_path or 'Закл' not in file_path:
                        continue
                    if '.xls' in file_path or '.pdf' in file_path or '.txt' in file_path:  # Пока работаем только с Word
                        continue

                    i = dance(i, well_folder_name)
                    research = extract_researches(file_path, well_folder_name)
                    all_researches.append(research)
                    area_number = get_area_number(file_path)

    write_to_excel(all_researches, os.path.join(output_folder_path, f'{area_number}.xlsx'))
