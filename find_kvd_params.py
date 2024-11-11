import datetime
import os
import re
import tkinter as tk
from tkinter import filedialog
import openpyxl
from docx import Document


def extract_info_from_table(table):
    data = []
    for row in table.rows:
        if len(row.cells) == 2:
            cell1 = row.cells[0].text
            cell2 = row.cells[1].text
            data.append([cell1, cell2])
    return data

def convert_doc_to_docx(doc_file, docx_file):
    # Используем antiword для конвертирования .doc в .docx
    subprocess.run(['antiword', doc_file, '-r', 'docx', '-w', docx_file])


def extract_info_from_docx(file_path, date, well_number):
    if file_path.endswith('.doc'):
        # Если файл .doc, то конвертируем его в .docx
        docx_file = file_path[:-4] + '.docx'
        os.system('antiword %s > %s' % (file_path, docx_file))
        file_path = docx_file

    doc = Document(file_path)

    table_data = []
    for table in doc.tables:
        table_data.extend(extract_info_from_table(table))
    write_to_excel(table_data, well_number, date)


def write_to_excel(table_data, well_number, date_of_kvd, turn_to_red=False):
    workbook_name = 'Результат по участку ' + well_number[0] + 'А.xlsx'
    date_of_kvd = date_of_kvd.strftime('%d.%m.%Y') if date_of_kvd else None  # Пока лучше не придумал для обработки None
    try:
        workbook = openpyxl.load_workbook(workbook_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    row = sheet.max_row + 1  # Находим первую пустую строку для добавления данных
    col = 1  # Номер колонки для номера скважины
    if turn_to_red:
        sheet.cell(row=row, column=col, value=well_number).fill = openpyxl.styles.PatternFill(start_color='FF0000',
                                                                                              fill_type='solid')
        workbook.save(workbook_name)
        return None
    sheet.cell(row=row, column=col, value=well_number)
    col = 2  # Номер колонки для даты
    sheet.cell(row=row, column=col, value=date_of_kvd)
    for table_row in table_data:
        col = 2
        for info in table_row:
            col += 1
            sheet.cell(row=row, column=col, value=info)
        row += 1

    workbook.save(workbook_name)


def find_well_folder(root_path, current_path):
    current_dir = current_path
    while current_dir != root_path:
        folder_name = os.path.basename(current_dir)
        match = re.search(r'\d+[АAА]\d+', folder_name, re.IGNORECASE)
        if match:
            return match.group()
        current_dir = os.path.dirname(current_dir)
    return None


def process_files_in_directory(directory_path):
    for root, dirs, files in os.walk(directory_path):
        well_folder_name = find_well_folder(directory_path, root)  # ищем номер скважины
        if well_folder_name:
            a = list(filter(lambda x: 'ГКИ' in x, dirs))
            if "ГКИ" in root:
                date = os.path.basename(root)  # Дата из названия папки
                try:
                    date = datetime.datetime.strptime(date, "%Y-%m-%d")  # Не помню, как конкретно оно там выглядит
                except ValueError:
                    date = None
                for file_name in files:
                    if file_name.startswith("Закл_КВД"):
                        file_path = os.path.join(root, file_name)
                        extract_info_from_docx(file_path, date,
                                               well_folder_name)
            # скважина может быть в root, когда в root нет папки ГКИ, хотя она есть в dirs. Поэтому надо проверять
            elif not list(filter(lambda x: 'ГКИ' in x, dirs)):
                write_to_excel(None, well_folder_name, None, True)


def browse_button():
    global folder_path
    folder_path = filedialog.askdirectory()
    process_files_in_directory(folder_path)


# Создание графического интерфейса
root = tk.Tk()
root.title("Выбор папки")
root.geometry("300x100")

browse_button = tk.Button(root, text="Выбрать папку", command=browse_button)
browse_button.pack(pady=20)

root.mainloop()
